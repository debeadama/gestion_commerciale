# views/clients_view.py
import os
from datetime import datetime as dt

from controllers.client_controller import ClientController
from PyQt6 import uic
from PyQt6.QtCore import Qt, QTimer
from PyQt6.QtGui import QColor, QFont
from PyQt6.QtWidgets import (
    QDialog,
    QDialogButtonBox,
    QFileDialog,
    QFormLayout,
    QFrame,
    QHBoxLayout,
    QLabel,
    QLineEdit,
    QMessageBox,
    QPushButton,
    QTableWidget,
    QTableWidgetItem,
    QVBoxLayout,
    QWidget,
)
from utils.pagination_widget import PaginationWidget





# Vue principale Clients



def _confirm(parent, title, message):
    box = QMessageBox(parent)
    box.setWindowTitle(title)
    box.setText(message)
    box.setIcon(QMessageBox.Icon.Question)
    btn_oui = box.addButton("Oui", QMessageBox.ButtonRole.YesRole)
    box.addButton("Non", QMessageBox.ButtonRole.NoRole)
    box.exec()
    return box.clickedButton() == btn_oui


class ClientsView(QWidget):

    def __init__(self, parent=None):
        super().__init__(parent)
        ui_path = os.path.join(os.path.dirname(__file__), 'ui', 'clients.ui')
        uic.loadUi(ui_path, self)
        self._connect_signals()
        self._apply_styles()
        self.load_clients()

    #   Signaux

    def _connect_signals(self):
        self.btn_add.clicked.connect(self._add_client)
        self.btn_edit.clicked.connect(self._edit_client)
        self.btn_delete.clicked.connect(self._delete_client)
        self.btn_history.clicked.connect(self._show_history)
        self.btn_search.clicked.connect(self._search)
        self.btn_reset.clicked.connect(self._reset)
        self.search_input.returnPressed.connect(self._search)
        # Debounce : attend 300ms apres la derniere frappe avant de chercher
        self._search_timer = QTimer(self)
        self._search_timer.setSingleShot(True)
        self._search_timer.setInterval(300)
        self._search_timer.timeout.connect(self._search)
        self.search_input.textChanged.connect(
            lambda: self._search_timer.start())
        self.clients_table.itemSelectionChanged.connect(
            self._on_selection_changed)
        self.clients_table.doubleClicked.connect(self._edit_client)
        if hasattr(self, 'btn_export_excel'):
            self.btn_export_excel.clicked.connect(self._export_excel)
        if hasattr(self, 'btn_export_pdf'):
            self.btn_export_pdf.clicked.connect(self._export_pdf)
        # Pagination — inserer sous le tableau
        self._all_clients = []
        self.pagination = PaginationWidget(page_size=50, parent=self)
        self.pagination.page_changed.connect(self._on_page_changed)
        # Trouver le layout parent du tableau et y ajouter la pagination
        table_parent = self.clients_table.parent()
        if table_parent and table_parent.layout():
            table_parent.layout().addWidget(self.pagination)

    #   Chargement

    def load_clients(self, search=None):
        self._all_clients = ClientController.get_all(search)
        count = len(self._all_clients) if self._all_clients else 0
        self.lbl_total.setText(f"Total : {count} client(s)")
        if hasattr(self, 'pagination'):
            self.pagination.reset(count)
            offset = self.pagination.current_offset()
            limit = self.pagination.current_limit()
        else:
            offset, limit = 0, 50
        self._populate_table(self._all_clients[offset:offset + limit])

    def _on_page_changed(self, page, offset, limit):
        self._populate_table(
            (self._all_clients or [])[offset:offset + limit])

    def _populate_table(self, clients):
        self.clients_table.setRowCount(0)
        self.clients_table.setColumnCount(7)
        self.clients_table.setHorizontalHeaderLabels(
            ["ID", "Nom", "Prenom", "Telephone", "Email", "Ville", "Code postal"])
        self.clients_table.setColumnHidden(0, True)
        self.clients_table.horizontalHeader().setStretchLastSection(True)
        self.clients_table.verticalHeader().setVisible(False)
        self.clients_table.setAlternatingRowColors(True)

        for row_idx, c in enumerate(clients or []):
            self.clients_table.insertRow(row_idx)
            vals = [
                str(c.get('id', '')),
                c.get('nom', ''),
                c.get('prenom', ''),
                c.get('telephone') or '',
                c.get('email') or '',
                c.get('ville') or '',
                c.get('code_postal') or '',
            ]
            for col_idx, val in enumerate(vals):
                item = QTableWidgetItem(val)
                item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                self.clients_table.setItem(row_idx, col_idx, item)

        self.clients_table.resizeColumnsToContents()
        if hasattr(self, 'status_bar'):
            self.status_bar.setText(
                f"{len(clients or [])} client(s) affiché(s)")

    #   Recherche

    def _reset(self):
        self.search_input.clear()
        self.load_clients()

    def _search(self):
        term = self.search_input.text().strip()
        self.load_clients(term if term else None)

    #  Selection

    def _on_selection_changed(self):
        has = len(self.clients_table.selectedItems()) > 0
        self.btn_edit.setEnabled(has)
        self.btn_delete.setEnabled(has)
        self.btn_history.setEnabled(has)

    def _get_selected_id(self):
        row = self.clients_table.currentRow()
        if row < 0:
            return None
        self.clients_table.setColumnHidden(0, False)
        id_item = self.clients_table.item(row, 0)
        self.clients_table.setColumnHidden(0, True)
        return int(id_item.text()) if id_item else None

    #   CRUD

    def _add_client(self):
        dialog = ClientFormDialog(parent=self)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            success, result = ClientController.create(dialog.get_data())
            if success:
                self.load_clients()
                QMessageBox.information(
                    self, "Succes", "Client ajoute avec succès.")
            else:
                QMessageBox.critical(self, "Erreur", result)

    def _edit_client(self):
        client_id = self._get_selected_id()
        if not client_id:
            return
        client = ClientController.get_by_id(client_id)
        if not client:
            return
        dialog = ClientFormDialog(client_data=client, parent=self)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            success, result = ClientController.update(
                client_id, dialog.get_data())
            if success:
                self.load_clients()
                QMessageBox.information(
                    self, "Succes", "Client modifie avec succès.")
            else:
                QMessageBox.critical(self, "Erreur", result)

    def _delete_client(self):
        client_id = self._get_selected_id()
        if not client_id:
            return
        row = self.clients_table.currentRow()
        nom = self.clients_table.item(row, 1).text()
        prenom = self.clients_table.item(row, 2).text()
        if _confirm(
            self,
            "Confirmation",
                f"Supprimér le client {nom} {prenom} ?"):
            success, message = ClientController.delete(client_id)
            if success:
                self.load_clients()
                QMessageBox.information(self, "Succes", message)
            else:
                QMessageBox.warning(self, "Impossible", message)

    def _show_history(self):
        client_id = self._get_selected_id()
        if not client_id:
            return
        row = self.clients_table.currentRow()
        nom = self.clients_table.item(row, 1).text()
        prenom = self.clients_table.item(row, 2).text()
        history = ClientController.get_purchase_history(client_id)
        stats = ClientController.get_stats(client_id)
        PurchaseHistoryDialog(
            f"{nom} {prenom}",
            history,
            stats,
            parent=self).exec()

    #  Export Excel

    def _export_excel(self):
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill

        filepath, _ = QFileDialog.getSaveFileName(
            self, "Exporter Excel",
            f"clients_{dt.now().strftime('%Y%m%d')}.xlsx",
            "Excel (*.xlsx)")
        if not filepath:
            return
        try:
            clients = ClientController.get_export_data()
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Clients"
            headers = ["Nom", "Prenom", "Telephone", "Email",
                       "Ville", "CA Total", "Dernière visite"]
            dark = PatternFill("solid", fgColor="1e293b")
            light = PatternFill("solid", fgColor="f1f5f9")

            for ci, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=ci, value=h)
                cell.fill = dark
                cell.font = Font(color="FFFFFF", bold=True, size=10)
                cell.alignment = Alignment(
                    horizontal="center", vertical="center")
            ws.row_dimensions[1].height = 22

            for ri, row in enumerate(clients or [], 2):
                fill = light if ri % 2 == 0 else PatternFill()
                row_vals = [
                    row.get("nom", ""),
                    row.get("prenom", ""),
                    row.get("telephone") or "",
                    row.get("email") or "",
                    row.get("ville") or "",
                    f"{float(row.get('ca_total', 0)):.2f}",
                    row.get("derniere_visite") or "N/A",
                ]
                for ci, val in enumerate(row_vals, 1):
                    cell = ws.cell(row=ri, column=ci, value=val)
                    cell.fill = fill
                    cell.alignment = Alignment(horizontal="center")

            for col_cells in ws.columns:
                ws.column_dimensions[col_cells[0].column_letter].width = 20

            wb.save(filepath)
            QMessageBox.information(self, "Succès",
                                    f"Export Excel reussi :\n{filepath}")
        except Exception as e:
            QMessageBox.critical(self, "Erreur export Excel", str(e))

    #  Export PDF

    def _export_pdf(self):
        from datetime import datetime as _dt

        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.platypus import (
            HRFlowable,
            Paragraph,
            SimpleDocTemplate,
            Spacer,
            Table,
            TableStyle,
        )

        filepath, _ = QFileDialog.getSaveFileName(
            self, "Exporter PDF",
            f"clients_{dt.now().strftime('%Y%m%d')}.pdf",
            "PDF (*.pdf)")
        if not filepath:
            return
        try:
            clients = ClientController.get_export_data()
            DARK = colors.HexColor("#1e293b")
            BLUE = colors.HexColor("#1976D2")
            LIGHT = colors.HexColor("#f1f5f9")
            GRAY = colors.HexColor("#64748b")

            doc = SimpleDocTemplate(
                filepath, pagesize=landscape(A4),
                rightMargin=1.5 * cm, leftMargin=1.5 * cm,
                topMargin=1.5 * cm, bottomMargin=1.5 * cm)

            story = []

            #  En-tête entreprise
            from utils.pdf_generator import get_company_info
            params = get_company_info()
            nom_entreprise = params.get('nom_entreprise', '')
            adresse = params.get('adresse', '')
            telephone = params.get('telephone', '')
            email_ent = params.get('email', '')

            # Ligne entreprise
            story.append(Spacer(1, 0.3 * cm))
            story.append(Paragraph(
                nom_entreprise,
                ParagraphStyle("Ent", fontSize=13, fontName="Helvetica-Bold",
                               textColor=BLUE, spaceAfter=4)))
            if adresse or telephone or email_ent:
                contact = " | ".join(
                    filter(None, [adresse, telephone, email_ent]))
                story.append(Paragraph(
                    contact,
                    ParagraphStyle("Contact", fontSize=8, fontName="Helvetica",
                                   textColor=GRAY, spaceAfter=6)))
            story.append(Spacer(1, 0.4 * cm))
            story.append(HRFlowable(width="100%", thickness=1.5,
                                    color=BLUE, spaceAfter=0.5 * cm))

            #  Titre du document
            story.append(Paragraph(
                "Liste des Clients",
                ParagraphStyle("T", fontSize=16, fontName="Helvetica-Bold",
                               textColor=DARK, spaceAfter=8)))
            story.append(Paragraph(
                f"Exportée le {dt.now().strftime('%d/%m/%Y à %H:%M')} "
                f"— {len(clients or [])} client(s)",
                ParagraphStyle("S", fontSize=8, fontName="Helvetica",
                               textColor=GRAY, spaceAfter=6)))
            story.append(Spacer(1, 0.6 * cm))

            #  Tableau
            headers = ["Nom", "Prénom", "Téléphone",
                       "Email", "Ville", "CA Total", "Dernière visite"]
            data = [headers]
            for row in (clients or []):
                # Formater derniere_visite en Python
                dv = row.get("derniere_visite")
                if dv:
                    try:
                        if hasattr(dv, 'strftime'):
                            dv_str = dv.strftime('%d/%m/%Y')
                        else:
                            dv_str = _dt.strptime(
                                str(dv)[:10], '%Y-%m-%d').strftime('%d/%m/%Y')
                    except Exception:
                        dv_str = str(dv)[:10]
                else:
                    dv_str = "N/A"
                data.append([
                    row.get("nom", ""),
                    row.get("prenom", "") or "",
                    row.get("telephone") or "",
                    row.get("email") or "",
                    row.get("ville") or "",
                    f"{float(row.get('ca_total', 0)):,.0f}",
                    dv_str,
                ])

            W = landscape(A4)[0] - 3 * cm
            col_w = [
                W * 0.12,
                W * 0.12,
                W * 0.15,
                W * 0.22,
                W * 0.12,
                W * 0.13,
                W * 0.14]
            t = Table(data, colWidths=col_w, repeatRows=1)
            t.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), DARK),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, LIGHT]),
                ("GRID", (0, 0), (-1, -1), 0.3,
                 colors.HexColor("#e2e8f0")),
                ("LINEBELOW", (0, 0), (-1, 0), 1.5, BLUE),
            ]))
            story.append(t)
            doc.build(story)
            QMessageBox.information(self, "Succes",
                                    f"Export PDF reussi :\n{filepath}")
        except Exception as e:
            QMessageBox.critical(self, "Erreur export PDF", str(e))

    #  Styles

    def _apply_styles(self):
        self.setStyleSheet("""
            QLabel#lbl_titre {
                font-size: 18px; font-weight: bold; color: #1e293b;
            }
            QLabel#lbl_total  { font-size: 11px; color: #94a3b8; }
            QLabel#status_bar { font-size: 11px; color: #94a3b8; padding: 2px 0; }
            QPushButton#btn_export_excel {
                background-color: #dcfce7; color: #16a34a;
            }
            QPushButton#btn_export_excel:hover { background-color: #bbf7d0; }
            QPushButton#btn_export_pdf {
                background-color: #fee2e2; color: #dc2626;
            }
            QPushButton#btn_export_pdf:hover { background-color: #fecaca; }
        """)


# Dialogue Formulaire Client


class ClientFormDialog(QDialog):

    def __init__(self, client_data=None, parent=None):
        super().__init__(parent)
        is_edit = client_data is not None
        self.setWindowTitle(
            "Modifier le client" if is_edit else "Nouveau client")
        self.setMinimumWidth(450)
        self.setModal(True)
        self._build_ui()
        if is_edit:
            self._fill_form(client_data)

    def _build_ui(self):
        layout = QVBoxLayout(self)
        layout.setSpacing(12)
        layout.setContentsMargins(20, 20, 20, 20)
        form = QFormLayout()
        form.setSpacing(10)
        self.nom_input = QLineEdit()
        self.prenom_input = QLineEdit()
        self.telephone_input = QLineEdit()
        self.email_input = QLineEdit()
        self.adresse_input = QLineEdit()
        self.ville_input = QLineEdit()
        self.postal_input = QLineEdit()
        for w in [self.nom_input, self.prenom_input, self.telephone_input,
                  self.email_input, self.adresse_input,
                  self.ville_input, self.postal_input]:
            w.setMinimumHeight(34)
        form.addRow("Nom *", self.nom_input)
        form.addRow("Prenom *", self.prenom_input)
        form.addRow("Telephone", self.telephone_input)
        form.addRow("Email", self.email_input)
        form.addRow("Adresse", self.adresse_input)
        form.addRow("Ville", self.ville_input)
        form.addRow("Code postal", self.postal_input)
        layout.addLayout(form)
        buttons = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Save |
            QDialogButtonBox.StandardButton.Cancel)
        buttons.button(
            QDialogButtonBox.StandardButton.Save).setText("Enregistrer")
        buttons.button(
            QDialogButtonBox.StandardButton.Cancel).setText("Annuler")
        buttons.accepted.connect(self._validate)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)

    def _fill_form(self, data):
        self.nom_input.setText(data.get('nom', ''))
        self.prenom_input.setText(data.get('prenom', ''))
        self.telephone_input.setText(data.get('telephone') or '')
        self.email_input.setText(data.get('email') or '')
        self.adresse_input.setText(data.get('adresse') or '')
        self.ville_input.setText(data.get('ville') or '')
        self.postal_input.setText(data.get('code_postal') or '')

    def _validate(self):
        if not self.nom_input.text().strip():
            QMessageBox.warning(
                self, "Champ requis", "Le nom est obligatoire.")
            self.nom_input.setFocus()
            return
        if not self.prenom_input.text().strip():
            QMessageBox.warning(
                self,
                "Champ requis",
                "Le prenom est obligatoire.")
            self.prenom_input.setFocus()
            return
        self.accept()

    def get_data(self):
        return {
            'nom': self.nom_input.text().strip(),
            'prenom': self.prenom_input.text().strip(),
            'telephone': self.telephone_input.text().strip(),
            'email': self.email_input.text().strip(),
            'adresse': self.adresse_input.text().strip(),
            'ville': self.ville_input.text().strip(),
            'code_postal': self.postal_input.text().strip(),
        }


# Dialogue Historique des achats


class PurchaseHistoryDialog(QDialog):

    def __init__(self, client_name, history, stats=None, parent=None):
        super().__init__(parent)
        self.setWindowTitle(f"Historique - {client_name}")
        self.setMinimumSize(750, 500)
        self._build_ui(client_name, history, stats or {})

    def _build_ui(self, client_name, history, stats):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(15, 15, 15, 15)
        layout.setSpacing(10)

        lbl = QLabel(f"Historique des achats : {client_name}")
        lbl.setFont(QFont("Arial", 13, QFont.Weight.Bold))
        layout.addWidget(lbl)

        if stats:
            stats_frame = QFrame()
            stats_frame.setStyleSheet(
                "QFrame { background: #f8fafc; border: 1px solid #e2e8f0;"
                " border-radius: 6px; }")
            sf = QHBoxLayout(stats_frame)
            sf.setContentsMargins(20, 10, 20, 10)
            sf.setSpacing(30)
            for i, (label, val) in enumerate([
                ("Nb achats", str(stats.get("nb_achats", 0))),
                ("CA total", f"{float(stats.get('ca_total', 0)):.2f}"),
                ("Dernière visite", stats.get("derniere_visite") or "N/A"),
            ]):
                block = QVBoxLayout()
                lbl_l = QLabel(label)
                lbl_l.setStyleSheet("color: #94a3b8; font-size: 10px;")
                lbl_v = QLabel(val)
                lbl_v.setStyleSheet(
                    "color: #1976D2; font-size: 15px; font-weight: bold;")
                block.addWidget(lbl_l)
                block.addWidget(lbl_v)
                sf.addLayout(block)
                if i < 2:
                    sep = QFrame()
                    sep.setFrameShape(QFrame.Shape.VLine)
                    sep.setStyleSheet("color: #e2e8f0;")
                    sf.addWidget(sep)
            layout.addWidget(stats_frame)

        table = QTableWidget()
        table.setColumnCount(6)
        table.setHorizontalHeaderLabels(
            ["N Facture", "Date", "Total", "Paye", "Reste", "Statut"])
        table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        table.setAlternatingRowColors(True)
        table.verticalHeader().setVisible(False)
        table.horizontalHeader().setStretchLastSection(True)

        statut_colors = {
            'payee': '#16a34a',
            'en_cours': '#1976D2',
            'partielle': '#d97706',
            'annulee': '#dc2626',
        }
        for row_idx, v in enumerate(history):
            table.insertRow(row_idx)
            statut = v.get('statut', '')
            vals = [
                v.get('numero_facture', ''),
                str(v.get('date_vente', ''))[:16],
                f"{float(v.get('montant_total', 0)):.2f}",
                f"{float(v.get('montant_paye', 0)):.2f}",
                f"{float(v.get('montant_reste', 0)):.2f}",
                statut.upper(),
            ]
            for col_idx, val in enumerate(vals):
                item = QTableWidgetItem(val)
                item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                if col_idx == 5:
                    item.setForeground(
                        QColor(statut_colors.get(statut, '#64748b')))
                    f = item.font()
                    f.setBold(True)
                    item.setFont(f)
                table.setItem(row_idx, col_idx, item)

        if not history:
            table.insertRow(0)
            item = QTableWidgetItem("Aucun achat enregistre pour ce client.")
            item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            item.setForeground(QColor('#94a3b8'))
            table.setItem(0, 0, item)
            table.setSpan(0, 0, 1, 6)

        layout.addWidget(table)

        btn_close = QPushButton("Fermer")
        btn_close.setFixedWidth(100)
        btn_close.setStyleSheet(
            "background-color: #64748b; color: white;"
            " border-radius: 5px; padding: 6px;")
        btn_close.clicked.connect(self.accept)
        h = QHBoxLayout()
        h.addStretch()
        h.addWidget(btn_close)
        layout.addLayout(h)
