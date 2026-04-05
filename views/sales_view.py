# views/sales_view.py
import os

from controllers.auth_controller import SessionManager
from controllers.client_controller import ClientController
from controllers.product_controller import ProductController
from controllers.sale_controller import SaleController
from PyQt6 import uic
from PyQt6.QtCore import QDate, Qt, QTimer
from PyQt6.QtGui import QColor, QFont
from PyQt6.QtWidgets import (
    QComboBox,
    QDateEdit,
    QDialog,
    QDialogButtonBox,
    QDoubleSpinBox,
    QFileDialog,
    QFormLayout,
    QFrame,
    QHBoxLayout,
    QLabel,
    QLineEdit,
    QMessageBox,
    QPushButton,
    QSpinBox,
    QTableWidget,
    QTableWidgetItem,
    QVBoxLayout,
    QWidget,
)
from utils.pagination_widget import PaginationWidget


def _confirm(parent, title, message):
    box = QMessageBox(parent)
    box.setWindowTitle(title)
    box.setText(message)
    box.setIcon(QMessageBox.Icon.Question)
    btn_oui = box.addButton("Oui", QMessageBox.ButtonRole.YesRole)
    box.addButton("Non", QMessageBox.ButtonRole.NoRole)
    box.exec()
    return box.clickedButton() == btn_oui


class SalesView(QWidget):
    """Module complet de gestion des ventes."""

    def __init__(self, parent=None):
        super().__init__(parent)
        ui_path = os.path.join(os.path.dirname(__file__), 'ui', 'sales.ui')
        uic.loadUi(ui_path, self)
        self._connect_signals()
        self._apply_styles()
        self.load_sales()
        if hasattr(self, "btn_validate"):
            can_see = SessionManager.has_permission("view_reports")
            self.btn_validate.setVisible(can_see)
        if hasattr(self, "btn_export_excel"):
            self.btn_export_excel.clicked.connect(self._export_excel)

    # ----------------------------------------------------------
    # Signaux
    # ----------------------------------------------------------

    def _connect_signals(self):
        self.btn_add.clicked.connect(self._add_sale)
        self.btn_view.clicked.connect(self._view_sale)
        self.btn_pdf.clicked.connect(self._print_pdf)
        self.btn_cancel.clicked.connect(self._cancel_sale)
        self.btn_validate.clicked.connect(self._validate_sale)
        self.btn_search.clicked.connect(self._search)
        self.btn_reset.clicked.connect(self._reset)
        self.search_input.returnPressed.connect(self._search)
        self.filter_statut.currentIndexChanged.connect(self._search)
        # Debounce : attend 300ms apres la derniere frappe avant de chercher
        self._search_timer = QTimer(self)
        self._search_timer.setSingleShot(True)
        self._search_timer.setInterval(300)
        self._search_timer.timeout.connect(self._search)
        self.search_input.textChanged.connect(
            lambda: self._search_timer.start())
        self.sales_table.itemSelectionChanged.connect(
            self._on_selection_changed)
        self.sales_table.doubleClicked.connect(self._view_sale)
        # Pagination
        self._all_sales = []
        self.pagination = PaginationWidget(page_size=50, parent=self)
        self.pagination.page_changed.connect(self._on_page_changed)
        table_parent = self.sales_table.parent()
        if table_parent and table_parent.layout():
            table_parent.layout().addWidget(self.pagination)

    # ----------------------------------------------------------
    # Chargement
    # ----------------------------------------------------------

    def load_sales(self, search=None, statut=None):
        self._all_sales = SaleController.get_all(search, statut)
        count = len(self._all_sales)
        self.lbl_total.setText(f"Total : {count} vente(s)")
        self.status_bar.setText(f"{count} vente(s) affichée(s)")
        self.pagination.reset(count)
        offset = self.pagination.current_offset()
        limit = self.pagination.current_limit()
        self._populate_table(self._all_sales[offset:offset + limit])

    def _on_page_changed(self, page, offset, limit):
        self._populate_table(
            (self._all_sales or [])[offset:offset + limit])

    @staticmethod
    def _format_date(date_val):
        """Convertit une date MySQL en format jj-mm-aaaa hh:mm:ss."""
        s = str(date_val)
        try:
            from datetime import datetime

            # Formats possibles : datetime object, 'YYYY-MM-DD HH:MM:SS',
            # 'YYYY-MM-DD HH:MM'
            if hasattr(date_val, 'strftime'):
                return date_val.strftime('%d-%m-%Y %H:%M:%S')
            s = s.strip()
            for fmt in ('%Y-%m-%d %H:%M:%S', '%Y-%m-%d %H:%M', '%Y-%m-%d'):
                try:
                    return datetime.strptime(
                        s[:19], fmt).strftime('%d-%m-%Y %H:%M:%S')
                except ValueError:
                    continue
        except Exception:
            pass
        return s[:19]

    def _populate_table(self, sales):
        self.sales_table.setRowCount(0)
        statut_colors = {
            'payee': '#16a34a',
            'partielle': '#d97706',
            'en_cours': '#1976D2',
            'annulee': '#dc2626',
        }
        for row_idx, s in enumerate(sales):
            self.sales_table.insertRow(row_idx)
            statut = s.get('statut', '')
            values = [
                str(s['id']),
                s.get('numero_facture', ''),
                s.get('client', ''),
                s.get('vendeur', ''),
                self._format_date(s.get("date_vente", "")),
                f"{float(s.get('montant_total', 0)):.2f}",
                f"{float(s.get('montant_paye', 0)):.2f}",
                f"{float(s.get('montant_reste', 0)):.2f}",
                statut.upper(),
            ]
            for col, val in enumerate(values):
                item = QTableWidgetItem(val)
                item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                if col == 8:
                    color = statut_colors.get(statut, '#64748b')
                    item.setForeground(QColor(color))
                    font = item.font()
                    font.setBold(True)
                    item.setFont(font)
                self.sales_table.setItem(row_idx, col, item)
        self.sales_table.resizeColumnsToContents()
        self.sales_table.setColumnHidden(0, True)

    # ----------------------------------------------------------
    # Actions
    # ----------------------------------------------------------

    def _reset(self):
        self.search_input.clear()
        self.filter_statut.setCurrentIndex(0)
        self.load_sales()

    def _search(self):
        term = self.search_input.text().strip() or None
        statut_map = {0: None, 1: 'en_cours', 2: 'payee',
                      3: 'partielle', 4: 'annulee'}
        statut = statut_map.get(self.filter_statut.currentIndex())
        self.load_sales(term, statut)

    def _on_selection_changed(self):
        has = len(self.sales_table.selectedItems()) > 0
        self.btn_view.setEnabled(has)
        self.btn_pdf.setEnabled(has)
        self.btn_cancel.setEnabled(has)
        if has and hasattr(self, "btn_validate"):
            statut = self._get_selected_statut()
            can_validate = (
                statut in ("en_cours", "partielle") and
                SessionManager.has_permission("view_reports"))
            self.btn_validate.setEnabled(can_validate)
        elif hasattr(self, "btn_validate"):
            self.btn_validate.setEnabled(False)

    def _get_selected_id(self):
        row = self.sales_table.currentRow()
        if row < 0:
            return None
        self.sales_table.setColumnHidden(0, False)
        id_item = self.sales_table.item(row, 0)
        self.sales_table.setColumnHidden(0, True)
        return int(id_item.text()) if id_item else None

    def _get_selected_statut(self):
        row = self.sales_table.currentRow()
        if row < 0:
            return None
        item = self.sales_table.item(row, 8)
        return item.text().lower() if item else None

    def _validate_sale(self):
        sale_id = self._get_selected_id()
        if not sale_id:
            return
        vente = SaleController.get_by_id(sale_id)
        if not vente:
            return
        if not _confirm(
            self,
            "Confirmer la validation",
            f"Valider la vente {
                vente.get(
                    'numero_facture',
                '')} ?\n\n" "Cette action marquera la vente comme PAYEE."):
            return
        success, message = SaleController.validate(sale_id)
        if success:
            self.load_sales()
            QMessageBox.information(self, "Succes", message)
        else:
            QMessageBox.warning(self, "Impossible", message)

    def _add_sale(self):
        dialog = SaleFormDialog(parent=self)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            data = dialog.get_data()
            success, result = SaleController.create(
                client_id=data['client_id'],
                panier=data['panier'],
                montant_paye=data['montant_paye'],
                notes=data['notes'],
            )
            if success:
                self.load_sales()
                vente = SaleController.get_by_id(result)
                numero = vente.get('numero_facture', '') if vente else ''
                self.status_bar.setText(f"Vente creee : {numero}")
                QMessageBox.information(
                    self, "Succes",
                    f"Vente enregistree avec succès !\nN Facture : {numero}")
            else:
                QMessageBox.critical(self, "Erreur", result)

    def _view_sale(self):
        sale_id = self._get_selected_id()
        if not sale_id:
            return
        vente = SaleController.get_by_id(sale_id)
        details = SaleController.get_details(sale_id)
        dialog = SaleDetailDialog(vente, details, parent=self)
        dialog.exec()
        self.load_sales()

    def _print_pdf(self):
        sale_id = self._get_selected_id()
        if not sale_id:
            return
        vente = SaleController.get_by_id(sale_id)
        details = SaleController.get_details(sale_id)
        try:
            from utils.pdf_generator import generate_invoice_pdf
            filepath = generate_invoice_pdf(vente, details)
            QMessageBox.information(
                self, "PDF genere",
                f"Facture generee avec succès !\n{filepath}")
        except Exception as e:
            QMessageBox.critical(self, "Erreur PDF", str(e))

    def _cancel_sale(self):
        sale_id = self._get_selected_id()
        if not sale_id:
            return
        row = self.sales_table.currentRow()
        numero = self.sales_table.item(row, 1).text()
        if _confirm(
            self,
            "Confirmation",
                f"Annuler la vente {numero} ?\nLe stock sera remis a jour."):
            success, message = SaleController.cancel(sale_id)
            if success:
                self.load_sales()
                self.status_bar.setText(f"Vente {numero} annulee.")
            else:
                QMessageBox.warning(self, "Impossible", message)

    # ----------------------------------------------------------
    # Export Excel avec filtres par date
    # ----------------------------------------------------------

    def _export_excel(self):
        dialog = ExportExcelDialog(parent=self)
        if dialog.exec() != QDialog.DialogCode.Accepted:
            return
        date_debut, date_fin = dialog.get_dates()

        filepath, _ = QFileDialog.getSaveFileName(
            self, "Exporter Excel",
            f"ventes_{date_debut}_{date_fin}.xlsx",
            "Excel (*.xlsx)")
        if not filepath:
            return
        try:
            import openpyxl
            from openpyxl.styles import Alignment, Font, PatternFill

            sales = SaleController.get_all()
            # Filtrer par date
            sales = [s for s in sales if date_debut <= str(
                s.get('date_vente', ''))[:10] <= date_fin]

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Ventes"
            headers = ["N Facture", "Client", "Vendeur", "Date",
                       "Total", "Paye", "Reste", "Statut"]
            dark = PatternFill("solid", fgColor="1e293b")
            light = PatternFill("solid", fgColor="f1f5f9")

            for ci, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=ci, value=h)
                cell.fill = dark
                cell.font = Font(color="FFFFFF", bold=True, size=10)
                cell.alignment = Alignment(
                    horizontal="center", vertical="center")
            ws.row_dimensions[1].height = 22

            for ri, s in enumerate(sales, 2):
                fill = light if ri % 2 == 0 else PatternFill()
                row_vals = [
                    s.get('numero_facture', ''),
                    s.get('client', ''),
                    s.get('vendeur', ''),
                    self._format_date(s.get('date_vente', '')),
                    f"{float(s.get('montant_total', 0)):.2f}",
                    f"{float(s.get('montant_paye', 0)):.2f}",
                    f"{float(s.get('montant_reste', 0)):.2f}",
                    s.get('statut', '').upper(),
                ]
                for ci, val in enumerate(row_vals, 1):
                    cell = ws.cell(row=ri, column=ci, value=val)
                    cell.fill = fill
                    cell.alignment = Alignment(horizontal="center")

            for col_cells in ws.columns:
                ws.column_dimensions[col_cells[0].column_letter].width = 20

            wb.save(filepath)
            QMessageBox.information(self, "Succes",
                                    f"Export Excel reussi :\n{filepath}")
        except Exception as e:
            QMessageBox.critical(self, "Erreur export", str(e))

    # ----------------------------------------------------------
    # Styles
    # ----------------------------------------------------------

    def _apply_styles(self):
        self.setStyleSheet("""
            QLabel#lbl_titre {
                font-size: 18px; font-weight: bold; color: #1e293b;
            }
            QLabel#lbl_total  { font-size: 11px; color: #94a3b8; }
            QLabel#status_bar { font-size: 11px; color: #94a3b8; padding: 2px 0; }
            QPushButton#btn_export_excel {
                background-color: #dcfce7; color: #16a34a;
            }
            QPushButton#btn_export_excel:hover { background-color: #bbf7d0; }
        """)


# ──────────────────────────────────────────────────────────────
# Dialogue filtre dates pour export Excel
# ──────────────────────────────────────────────────────────────

class ExportExcelDialog(QDialog):

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Exporter les ventes")
        self.setMinimumWidth(350)
        self.setModal(True)
        layout = QVBoxLayout(self)
        layout.setSpacing(12)
        layout.setContentsMargins(20, 20, 20, 20)

        layout.addWidget(QLabel("Choisissez la période à exporter :"))

        form = QFormLayout()
        self.date_debut = QDateEdit()
        self.date_debut.setCalendarPopup(True)
        self.date_debut.setDate(QDate.currentDate().addMonths(-1))
        self.date_debut.setDisplayFormat("dd/MM/yyyy")

        self.date_fin = QDateEdit()
        self.date_fin.setCalendarPopup(True)
        self.date_fin.setDate(QDate.currentDate())
        self.date_fin.setDisplayFormat("dd/MM/yyyy")

        form.addRow("Date début :", self.date_debut)
        form.addRow("Date fin :", self.date_fin)
        layout.addLayout(form)

        buttons = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Ok |
            QDialogButtonBox.StandardButton.Cancel)
        buttons.button(QDialogButtonBox.StandardButton.Ok).setText("Exporter")
        buttons.button(
            QDialogButtonBox.StandardButton.Cancel).setText("Annuler")
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)

    def get_dates(self):
        return (
            self.date_debut.date().toString("yyyy-MM-dd"),
            self.date_fin.date().toString("yyyy-MM-dd"),
        )


# ──────────────────────────────────────────────────────────────
# Formulaire Nouvelle Vente (avec remise et TVA)
# ──────────────────────────────────────────────────────────────

class SaleFormDialog(QDialog):

    TVA_RATE = 0.18  # 18% - modifiable selon le pays

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Nouvelle vente")
        self.setMinimumSize(800, 650)
        self.setModal(True)
        self.panier = []
        self._build_ui()
        self._load_clients()
        self._load_products()

    def _build_ui(self):
        layout = QVBoxLayout(self)
        layout.setSpacing(10)
        layout.setContentsMargins(20, 20, 20, 20)

        lbl = QLabel("Nouvelle Vente")
        lbl.setFont(QFont("Arial", 14, QFont.Weight.Bold))
        layout.addWidget(lbl)

        line = QFrame()
        line.setFrameShape(QFrame.Shape.HLine)
        layout.addWidget(line)

        # Client + Notes
        row1 = QHBoxLayout()
        form1 = QFormLayout()
        self.client_combo = QComboBox()
        self.client_combo.setMinimumHeight(34)
        self.client_combo.setMinimumWidth(280)
        # Bouton création rapide client
        self.btn_new_client = QPushButton("+ Nouveau client")
        self.btn_new_client.setFixedHeight(34)
        self.btn_new_client.setStyleSheet("""
            QPushButton {
                background: #e3f2fd; color: #1976D2;
                border: 1px solid #90caf9; border-radius: 5px;
                padding: 0 10px; font-size: 12px; font-weight: bold;
            }
            QPushButton:hover { background: #bbdefb; }
        """)
        self.btn_new_client.clicked.connect(self._create_client_quick)
        client_row = QHBoxLayout()
        client_row.setSpacing(6)
        client_row.addWidget(self.client_combo)
        client_row.addWidget(self.btn_new_client)
        form1.addRow("Client *", client_row)
        row1.addLayout(form1)
        row1.addStretch()
        form2 = QFormLayout()
        self.notes_input = QLineEdit()
        self.notes_input.setMinimumHeight(34)
        self.notes_input.setMinimumWidth(200)
        self.notes_input.setPlaceholderText("Notes optionnelles...")
        form2.addRow("Notes", self.notes_input)
        row1.addLayout(form2)
        layout.addLayout(row1)

        # Ajout produit
        sep1 = QLabel("Ajouter un produit au panier")
        sep1.setStyleSheet(
            "font-weight: bold; color: #475569; margin-top: 5px;")
        layout.addWidget(sep1)

        product_bar = QHBoxLayout()
        product_bar.setSpacing(8)
        self.product_combo = QComboBox()
        self.product_combo.setMinimumHeight(34)
        self.product_combo.setMinimumWidth(220)
        self.qty_input = QSpinBox()
        self.qty_input.setMinimumHeight(34)
        self.qty_input.setMinimum(1)
        self.qty_input.setMaximum(9999)
        self.price_input = QDoubleSpinBox()
        self.price_input.setMinimumHeight(34)
        self.price_input.setMaximum(9999999)
        self.price_input.setDecimals(2)
        self.btn_add_line = QPushButton("Ajouter au panier")
        self.btn_add_line.setMinimumHeight(34)
        self.btn_add_line.setStyleSheet(
            "background-color:#1976D2;color:white;border:none;"
            "border-radius:5px;padding:5px 12px;")
        product_bar.addWidget(QLabel("Produit:"))
        product_bar.addWidget(self.product_combo)
        product_bar.addWidget(QLabel("Qte:"))
        product_bar.addWidget(self.qty_input)
        product_bar.addWidget(QLabel("Prix:"))
        product_bar.addWidget(self.price_input)
        product_bar.addWidget(self.btn_add_line)
        layout.addLayout(product_bar)

        # Tableau panier
        self.cart_table = QTableWidget()
        self.cart_table.setColumnCount(5)
        self.cart_table.setHorizontalHeaderLabels(
            ["Produit", "Prix unitaire", "Quantite", "Sous-total", ""])
        self.cart_table.setEditTriggers(
            QTableWidget.EditTrigger.NoEditTriggers)
        self.cart_table.setSelectionBehavior(
            QTableWidget.SelectionBehavior.SelectRows)
        self.cart_table.verticalHeader().setVisible(False)
        self.cart_table.horizontalHeader().setStretchLastSection(True)
        self.cart_table.setMaximumHeight(180)
        layout.addWidget(self.cart_table)

        # Remise + TVA + Totaux + Paiement
        bottom = QHBoxLayout()
        bottom.addStretch()

        form3 = QFormLayout()
        form3.setSpacing(8)

        # Remise
        remise_layout = QHBoxLayout()
        self.remise_input = QDoubleSpinBox()
        self.remise_input.setMinimumHeight(32)
        self.remise_input.setMaximum(9999999)
        self.remise_input.setDecimals(2)
        self.remise_type = QComboBox()
        self.remise_type.setMinimumHeight(32)
        self.remise_type.addItem("Montant fixe", "montant")
        self.remise_type.addItem("Pourcentage (%)", "pourcentage")
        remise_layout.addWidget(self.remise_input)
        remise_layout.addWidget(self.remise_type)

        # TVA
        self.tva_combo = QComboBox()
        self.tva_combo.setMinimumHeight(32)
        self.tva_combo.addItem("Sans TVA (0%)", 0.0)
        self.tva_combo.addItem("TVA 9%", 0.09)
        self.tva_combo.addItem("TVA 18%", 0.18)
        self.tva_combo.addItem("TVA 20%", 0.20)
        self.tva_combo.setCurrentIndex(2)  # TVA 18% par defaut (Cote d'Ivoire)

        self.lbl_sous_total = QLabel("0.00")
        self.lbl_remise_val = QLabel("0.00")
        self.lbl_tva_val = QLabel("0.00")
        self.lbl_total = QLabel("0.00")
        self.lbl_total.setStyleSheet(
            "font-size:16px;font-weight:bold;color:#1976D2;")

        self.montant_paye_input = QDoubleSpinBox()
        self.montant_paye_input.setMinimumHeight(34)
        self.montant_paye_input.setMaximum(9999999)
        self.montant_paye_input.setDecimals(2)

        self.lbl_reste = QLabel("0.00")
        self.lbl_reste.setStyleSheet(
            "font-size:13px;color:#dc2626;font-weight:bold;")

        form3.addRow("Sous-total HT :", self.lbl_sous_total)
        form3.addRow("Remise :", remise_layout)
        form3.addRow("Remise (valeur) :", self.lbl_remise_val)
        form3.addRow("TVA :", self.tva_combo)
        form3.addRow("TVA (valeur) :", self.lbl_tva_val)
        form3.addRow("Total TTC :", self.lbl_total)
        form3.addRow("Montant payé :", self.montant_paye_input)
        form3.addRow("Reste a payer :", self.lbl_reste)
        bottom.addLayout(form3)
        layout.addLayout(bottom)

        # Boutons
        buttons = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Save |
            QDialogButtonBox.StandardButton.Cancel)
        buttons.button(
            QDialogButtonBox.StandardButton.Save).setText("Enregistrer")
        buttons.button(
            QDialogButtonBox.StandardButton.Cancel).setText("Annuler")
        buttons.accepted.connect(self._validate)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)

        # Connexions
        self.btn_add_line.clicked.connect(self._add_to_cart)
        self.product_combo.currentIndexChanged.connect(
            self._on_product_selected)
        self.montant_paye_input.valueChanged.connect(self._update_reste)
        self.remise_input.valueChanged.connect(self._update_total)
        self.remise_type.currentIndexChanged.connect(self._update_total)
        self.tva_combo.currentIndexChanged.connect(self._update_total)

    def _load_clients(self):
        self.client_combo.clear()
        self.client_combo.addItem("-- Selectionnez un client --", None)
        for c in ClientController.get_all():
            self.client_combo.addItem(
                f"{c['nom']} {c['prenom']} - {c.get('telephone', '')}",
                c['id'])

    def _create_client_quick(self):
        """Ouvre un dialog de création rapide de client depuis la vente."""
        dialog = QuickClientDialog(self)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            data = dialog.get_data()
            ok, result = ClientController.create(data)
            if ok:
                # Recharger la liste et sélectionner le nouveau client
                self._load_clients()
                for i in range(self.client_combo.count()):
                    if self.client_combo.itemData(i) == result:
                        self.client_combo.setCurrentIndex(i)
                        break
                QMessageBox.information(
                    self, "Succès", f"Client {
                        data['nom']} {
                        data.get(
                            'prenom', '')} créé et sélectionné.")
            else:
                QMessageBox.warning(self, "Erreur", result)

    def _load_products(self):
        self.product_combo.clear()
        for p in ProductController.get_all():
            self.product_combo.addItem(
                f"{p['nom']} (Stock: {p['stock_actuel']})", p)

    def _on_product_selected(self):
        data = self.product_combo.currentData()
        if data:
            self.price_input.setValue(float(data.get('prix_vente', 0)))

    def _add_to_cart(self):
        product_data = self.product_combo.currentData()
        if not product_data:
            QMessageBox.warning(self, "Attention", "Selectionnez un produit.")
            return
        qty = self.qty_input.value()
        price = self.price_input.value()
        if price <= 0:
            QMessageBox.warning(self, "Attention",
                                "Le prix doit etre superieur a 0.")
            return
        if qty > product_data['stock_actuel']:
            QMessageBox.warning(
                self, "Stock insuffisant",
                f"Stock disponible : {product_data['stock_actuel']}")
            return
        for ligne in self.panier:
            if ligne['produit_id'] == product_data['id']:
                ligne['quantite'] += qty
                self._refresh_cart()
                self._update_total()
                return
        self.panier.append({
            'produit_id': product_data['id'],
            'nom': product_data['nom'],
            'quantite': qty,
            'prix_unitaire': price,
        })
        self._refresh_cart()
        self._update_total()

    def _refresh_cart(self):
        self.cart_table.setRowCount(0)
        for row_idx, ligne in enumerate(self.panier):
            self.cart_table.insertRow(row_idx)
            sous_total = ligne['quantite'] * ligne['prix_unitaire']
            vals = [
                ligne['nom'],
                f"{ligne['prix_unitaire']:.2f}",
                str(ligne['quantite']),
                f"{sous_total:.2f}",
            ]
            for col, val in enumerate(vals):
                item = QTableWidgetItem(val)
                item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                self.cart_table.setItem(row_idx, col, item)
            btn_del = QPushButton("X")
            btn_del.setStyleSheet(
                "background-color:#ef4444;color:white;border:none;"
                "border-radius:3px;padding:2px 8px;")
            btn_del.clicked.connect(lambda _, r=row_idx: self._remove_line(r))
            self.cart_table.setCellWidget(row_idx, 4, btn_del)

    def _remove_line(self, row_idx):
        if 0 <= row_idx < len(self.panier):
            self.panier.pop(row_idx)
            self._refresh_cart()
            self._update_total()

    def _update_total(self):
        """Calcule sous-total HT, remise, TVA et total TTC."""
        sous_total = sum(
            item_p['quantite'] * item_p['prix_unitaire'] for item_p in self.panier)
        self.lbl_sous_total.setText(f"{sous_total:.2f}")

        # Remise
        remise_val = self.remise_input.value()
        if self.remise_type.currentData() == 'pourcentage':
            remise = sous_total * (remise_val / 100)
        else:
            remise = min(remise_val, sous_total)
        self.lbl_remise_val.setText(f"{remise:.2f}")

        apres_remise = sous_total - remise

        # TVA
        tva_rate = self.tva_combo.currentData() or 0.0
        tva_val = apres_remise * tva_rate
        self.lbl_tva_val.setText(f"{tva_val:.2f}")

        total = apres_remise + tva_val
        self.lbl_total.setText(f"{total:.2f}")
        self.montant_paye_input.setMaximum(total)
        self._update_reste()

    def _update_reste(self):
        try:
            total = float(self.lbl_total.text())
        except ValueError:
            total = 0
        reste = total - self.montant_paye_input.value()
        self.lbl_reste.setText(f"{max(reste, 0):.2f}")

    def _validate(self):
        if self.client_combo.currentData() is None:
            QMessageBox.warning(self, "Client requis",
                                "Selectionnez un client.")
            return
        if not self.panier:
            QMessageBox.warning(self, "Panier vide",
                                "Ajoutez au moins un produit au panier.")
            return
        self.accept()

    def get_data(self):
        try:
            total = float(self.lbl_total.text())
        except ValueError:
            total = 0
        return {
            'client_id': self.client_combo.currentData(),
            'panier': self.panier,
            'montant_paye': self.montant_paye_input.value(),
            'notes': self.notes_input.text().strip(),
            'montant_total': total,
        }


# ──────────────────────────────────────────────────────────────
# Création rapide de client depuis la vente
# ──────────────────────────────────────────────────────────────

class QuickClientDialog(QDialog):
    """Dialog de création rapide d'un client depuis le formulaire de vente."""

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Nouveau client")
        self.setMinimumWidth(420)
        self.setModal(True)
        self._build_ui()

    def _build_ui(self):
        layout = QVBoxLayout(self)
        layout.setSpacing(12)
        layout.setContentsMargins(20, 20, 20, 20)

        title = QLabel("Créer un nouveau client")
        title.setStyleSheet(
            "font-size: 15px; font-weight: bold; color: #1e293b;")
        layout.addWidget(title)

        form = QFormLayout()
        form.setSpacing(10)

        self.nom_input = QLineEdit()
        self.nom_input.setPlaceholderText("Obligatoire")
        self.prenom_input = QLineEdit()
        self.prenom_input.setPlaceholderText("Optionnel")
        self.tel_input = QLineEdit()
        self.tel_input.setPlaceholderText("Ex: +225 07 00 00 00")
        self.email_input = QLineEdit()
        self.email_input.setPlaceholderText("Optionnel")
        self.ville_input = QLineEdit()
        self.ville_input.setPlaceholderText("Optionnel")

        for w in [self.nom_input, self.prenom_input, self.tel_input,
                  self.email_input, self.ville_input]:
            w.setMinimumHeight(32)

        form.addRow("Nom *", self.nom_input)
        form.addRow("Prénom", self.prenom_input)
        form.addRow("Téléphone", self.tel_input)
        form.addRow("Email", self.email_input)
        form.addRow("Ville", self.ville_input)
        layout.addLayout(form)

        # Boutons
        btn_layout = QHBoxLayout()
        btn_layout.addStretch()
        btn_cancel = QPushButton("Annuler")
        btn_cancel.setFixedHeight(34)
        btn_cancel.setStyleSheet(
            "background:#f1f5f9; color:#1e293b; border:1px solid #cbd5e1; border-radius:5px; padding:0 16px;")
        btn_cancel.clicked.connect(self.reject)

        btn_save = QPushButton("Créer le client")
        btn_save.setFixedHeight(34)
        btn_save.setStyleSheet(
            "background:#1976D2; color:white; border:none; border-radius:5px; padding:0 16px; font-weight:bold;")
        btn_save.clicked.connect(self._validate)

        btn_layout.addWidget(btn_cancel)
        btn_layout.addWidget(btn_save)
        layout.addLayout(btn_layout)

    def _validate(self):
        if not self.nom_input.text().strip():
            QMessageBox.warning(
                self,
                "Champ manquant",
                "Le nom du client est obligatoire.")
            self.nom_input.setFocus()
            return
        self.accept()

    def get_data(self) -> dict:
        return {
            'nom': self.nom_input.text().strip(),
            'prenom': self.prenom_input.text().strip(),
            'telephone': self.tel_input.text().strip() or None,
            'email': self.email_input.text().strip() or None,
            'ville': self.ville_input.text().strip() or None,
            'adresse': None,
            'code_postal': None,
            'notes': None,
        }


# ──────────────────────────────────────────────────────────────
# Detail d'une vente + Paiement
# ──────────────────────────────────────────────────────────────

class SaleDetailDialog(QDialog):

    def __init__(self, vente, details, parent=None):
        super().__init__(parent)
        self.vente = vente
        self.setWindowTitle(
            f"Détail vente - {vente.get('numero_facture', '')}")
        self.setMinimumSize(650, 500)
        self._build_ui(vente, details)

    def _build_ui(self, vente, details):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(10)

        info_layout = QHBoxLayout()
        left = QVBoxLayout()
        left.addWidget(
            QLabel(
                f"N Facture : {
                    vente.get(
                        'numero_facture',
                        '')}"))
        left.addWidget(QLabel(f"Client    : {vente.get('client_nom', '')}"))
        left.addWidget(QLabel(f"Vendeur   : {vente.get('vendeur', '')}"))
        left.addWidget(
            QLabel(
                f"Date      : {
                    SalesView._format_date(
                        vente.get(
                            'date_vente',
                            ''))}"))

        right = QVBoxLayout()
        right.addWidget(QLabel(
            f"Total     : {float(vente.get('montant_total', 0)):.2f}"))
        right.addWidget(QLabel(
            f"Payé      : {float(vente.get('montant_paye', 0)):.2f}"))
        lbl_reste = QLabel(
            f"Reste     : {float(vente.get('montant_reste', 0)):.2f}")
        lbl_reste.setStyleSheet("color: #dc2626; font-weight: bold;")
        right.addWidget(lbl_reste)
        statut_colors = {
            'payee': '#16a34a', 'partielle': '#d97706',
            'en_cours': '#1976D2', 'annulee': '#dc2626'}
        lbl_statut = QLabel(f"Statut    : {vente.get('statut', '').upper()}")
        lbl_statut.setStyleSheet(
            f"color: {statut_colors.get(vente.get('statut'), '#64748b')};"
            "font-weight: bold;")
        right.addWidget(lbl_statut)
        info_layout.addLayout(left)
        info_layout.addStretch()
        info_layout.addLayout(right)
        layout.addLayout(info_layout)

        line = QFrame()
        line.setFrameShape(QFrame.Shape.HLine)
        layout.addWidget(line)

        lbl = QLabel("Produits commandés")
        lbl.setFont(QFont("Arial", 12, QFont.Weight.Bold))
        layout.addWidget(lbl)

        table = QTableWidget()
        table.setColumnCount(4)
        table.setHorizontalHeaderLabels(
            ["Produit", "Prix unitaire", "Quantite", "Sous-total"])
        table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        table.setAlternatingRowColors(True)
        table.verticalHeader().setVisible(False)
        table.horizontalHeader().setStretchLastSection(True)
        table.setMaximumHeight(180)

        for row_idx, d in enumerate(details):
            table.insertRow(row_idx)
            vals = [
                d.get('produit', ''),
                f"{float(d.get('prix_unitaire', 0)):.2f}",
                str(d.get('quantite', 0)),
                f"{float(d.get('sous_total', 0)):.2f}",
            ]
            for col, val in enumerate(vals):
                item = QTableWidgetItem(val)
                item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                table.setItem(row_idx, col, item)
        layout.addWidget(table)

        # Paiement complementaire
        if vente.get('statut') not in ('payee', 'annulee'):
            line2 = QFrame()
            line2.setFrameShape(QFrame.Shape.HLine)
            layout.addWidget(line2)
            lbl2 = QLabel("Ajouter un paiement")
            lbl2.setFont(QFont("Arial", 12, QFont.Weight.Bold))
            layout.addWidget(lbl2)
            pay_layout = QHBoxLayout()
            self.pay_input = QDoubleSpinBox()
            self.pay_input.setMinimumHeight(36)
            self.pay_input.setMaximum(float(vente.get('montant_reste', 0)))
            self.pay_input.setValue(float(vente.get('montant_reste', 0)))
            self.pay_input.setDecimals(2)
            btn_pay = QPushButton("Valider le paiement")
            btn_pay.setMinimumHeight(36)
            btn_pay.setStyleSheet(
                "background-color:#1976D2;color:white;border:none;"
                "border-radius:5px;padding:5px 15px;font-weight:bold;")
            btn_pay.clicked.connect(self._add_payment)
            pay_layout.addWidget(QLabel("Montant :"))
            pay_layout.addWidget(self.pay_input)
            pay_layout.addWidget(btn_pay)
            pay_layout.addStretch()
            layout.addLayout(pay_layout)

        btn_close = QPushButton("Fermer")
        btn_close.setFixedWidth(100)
        btn_close.setStyleSheet(
            "background-color:#64748b;color:white;border:none;"
            "border-radius:5px;padding:6px;")
        btn_close.clicked.connect(self.accept)
        h = QHBoxLayout()
        h.addStretch()
        h.addWidget(btn_close)
        layout.addLayout(h)

    def _add_payment(self):
        montant = self.pay_input.value()
        if montant <= 0:
            QMessageBox.warning(self, "Attention",
                                "Le montant doit etre superieur a 0.")
            return
        success, message = SaleController.add_payment(
            self.vente['id'], montant)
        if success:
            QMessageBox.information(self, "Succes", message)
            self.accept()
        else:
            QMessageBox.warning(self, "Erreur", message)
