# views/stats_view.py
from datetime import datetime

from controllers.product_controller import ProductController
from controllers.sale_controller import SaleController
from controllers.user_controller import UserController
from PyQt6.QtCore import QDate, Qt
from PyQt6.QtGui import QColor, QFont
from PyQt6.QtWidgets import (
    QComboBox,
    QDateEdit,
    QFileDialog,
    QFrame,
    QHBoxLayout,
    QLabel,
    QMessageBox,
    QPushButton,
    QSizePolicy,
    QTableWidget,
    QTableWidgetItem,
    QTabWidget,
    QVBoxLayout,
    QWidget,
)


class StatsView(QWidget):
    """Module Statistiques et Rapports."""

    def __init__(self, parent=None):
        super().__init__(parent)
        self._build_ui()
        self._apply_styles()
        self.refresh()

    # ----------------------------------------------------------
    # Interface principale
    # ----------------------------------------------------------

    def _build_ui(self):
        main = QVBoxLayout(self)
        main.setSpacing(12)
        main.setContentsMargins(20, 15, 20, 15)

        header = QHBoxLayout()
        lbl = QLabel("Statistiques et Rapports")
        lbl.setFont(QFont("Arial", 18, QFont.Weight.Bold))
        lbl.setStyleSheet("color: #1e293b;")
        self.lbl_refresh = QLabel("")
        self.lbl_refresh.setStyleSheet("color: #94a3b8; font-size: 11px;")
        btn_refresh = QPushButton("Actualiser")
        btn_refresh.setFixedWidth(100)
        btn_refresh.setFixedHeight(32)
        btn_refresh.setObjectName("btn_refresh")
        btn_refresh.clicked.connect(self.refresh)
        header.addWidget(lbl)
        header.addStretch()
        header.addWidget(self.lbl_refresh)
        header.addWidget(btn_refresh)
        main.addLayout(header)

        sep = QFrame()
        sep.setFrameShape(QFrame.Shape.HLine)
        sep.setFrameShadow(QFrame.Shadow.Sunken)
        main.addWidget(sep)

        self.tabs = QTabWidget()
        self.tabs.setObjectName("main_tabs")
        main.addWidget(self.tabs)

        self.tab_stats = QWidget()
        self._build_stats_tab()
        self.tabs.addTab(self.tab_stats, "  Statistiques  ")

        self.tab_reports = QWidget()
        self._build_reports_tab()
        self.tabs.addTab(self.tab_reports, "  Rapports  ")

    # ----------------------------------------------------------
    # Onglet Statistiques
    # ----------------------------------------------------------

    def _build_stats_tab(self):
        layout = QVBoxLayout(self.tab_stats)
        layout.setSpacing(12)
        layout.setContentsMargins(0, 12, 0, 0)

        # Cartes CA
        ca_layout = QHBoxLayout()
        ca_layout.setSpacing(12)
        self.card_jour = self._make_card("CA Aujourd'hui", "0.00", "#1976D2")
        self.card_semaine = self._make_card(
            "CA Cette Semaine", "0.00", "#059669")
        self.card_mois = self._make_card("CA Ce Mois", "0.00", "#7c3aed")
        ca_layout.addWidget(self.card_jour['frame'])
        ca_layout.addWidget(self.card_semaine['frame'])
        ca_layout.addWidget(self.card_mois['frame'])
        layout.addLayout(ca_layout)

        # Tops + Stock critique
        bottom = QHBoxLayout()
        bottom.setSpacing(12)

        for attr, title, headers in [
            ('top_products_table', 'Top 5 Produits Vendus',
             ["Produit", "Qte vendue", "CA"]),
            ('top_clients_table', 'Top 5 Clients',
             ["Client", "Nb ventes", "CA total"]),
            ('stock_table', 'Etat Stock Critique',
             ["Produit", "Stock", "Min"]),
        ]:
            frame = QFrame()
            frame.setObjectName("stat_frame")
            fl = QVBoxLayout(frame)
            fl.setContentsMargins(15, 12, 15, 12)
            lbl = QLabel(title)
            lbl.setFont(QFont("Arial", 11, QFont.Weight.Bold))
            fl.addWidget(lbl)
            table = self._make_mini_table(headers)
            setattr(self, attr, table)
            fl.addWidget(table)
            bottom.addWidget(frame)

        layout.addLayout(bottom)

        # Graphique evolution CA
        graph_frame = QFrame()
        graph_frame.setObjectName("stat_frame")
        graph_layout = QVBoxLayout(graph_frame)
        graph_layout.setContentsMargins(15, 12, 15, 12)
        lbl_g = QLabel("Evolution du Chiffre d'Affaires (6 derniers mois)")
        lbl_g.setFont(QFont("Arial", 11, QFont.Weight.Bold))
        graph_layout.addWidget(lbl_g)
        self.graph_widget = BarChartWidget()
        self.graph_widget.setFixedHeight(180)
        graph_layout.addWidget(self.graph_widget)
        layout.addWidget(graph_frame)

    # ----------------------------------------------------------
    # Onglet Rapports
    # ----------------------------------------------------------

    def _build_reports_tab(self):
        layout = QVBoxLayout(self.tab_reports)
        layout.setSpacing(15)
        layout.setContentsMargins(0, 12, 0, 0)

        # Filtres
        filter_frame = QFrame()
        filter_frame.setObjectName("stat_frame")
        filter_layout = QHBoxLayout(filter_frame)
        filter_layout.setContentsMargins(15, 12, 15, 12)
        filter_layout.setSpacing(10)
        filter_layout.addWidget(QLabel("Période :"))
        self.date_debut = QDateEdit()
        self.date_debut.setCalendarPopup(True)
        self.date_debut.setDate(QDate.currentDate().addDays(-30))
        self.date_debut.setFixedHeight(34)
        self.date_fin = QDateEdit()
        self.date_fin.setCalendarPopup(True)
        self.date_fin.setDate(QDate.currentDate())
        self.date_fin.setFixedHeight(34)
        self.filter_vendeur = QComboBox()
        self.filter_vendeur.setFixedHeight(34)
        self.filter_vendeur.setFixedWidth(160)
        self.filter_vendeur.addItem("Tous les vendeurs", None)
        btn_apply = QPushButton("Appliquer")
        btn_apply.setFixedHeight(34)
        btn_apply.setObjectName("btn_primary")
        btn_apply.clicked.connect(self._load_report_data)
        filter_layout.addWidget(self.date_debut)
        filter_layout.addWidget(QLabel("au"))
        filter_layout.addWidget(self.date_fin)
        filter_layout.addWidget(QLabel("Vendeur :"))
        filter_layout.addWidget(self.filter_vendeur)
        filter_layout.addWidget(btn_apply)
        filter_layout.addStretch()
        layout.addWidget(filter_frame)

        # Tableaux rapports
        report_frame = QFrame()
        report_frame.setObjectName("stat_frame")
        report_layout = QVBoxLayout(report_frame)
        report_layout.setContentsMargins(15, 12, 15, 12)
        self.report_tabs = QTabWidget()
        self.report_tabs.setObjectName("report_tabs")

        # Rapport ventes
        rw = QWidget()
        rl = QVBoxLayout(rw)
        rl.setContentsMargins(0, 8, 0, 0)
        self.rpt_ventes_table = QTableWidget()
        self.rpt_ventes_table.setColumnCount(7)
        self.rpt_ventes_table.setHorizontalHeaderLabels(
            ["N Facture", "Date", "Client", "Vendeur",
             "Total", "Paye", "Statut"])
        self._style_table(self.rpt_ventes_table)
        rl.addWidget(self.rpt_ventes_table)
        self.lbl_total_ventes = QLabel("Total : 0.00 | Nb ventes : 0")
        self.lbl_total_ventes.setStyleSheet(
            "font-weight: bold; color: #1976D2; padding: 5px;")
        rl.addWidget(self.lbl_total_ventes)
        self.report_tabs.addTab(rw, "Ventes par periode")

        # Rapport impayes
        ri = QWidget()
        ril = QVBoxLayout(ri)
        ril.setContentsMargins(0, 8, 0, 0)
        self.rpt_impaye_table = QTableWidget()
        self.rpt_impaye_table.setColumnCount(6)
        self.rpt_impaye_table.setHorizontalHeaderLabels(
            ["N Facture", "Date", "Client", "Total", "Paye", "Reste"])
        self._style_table(self.rpt_impaye_table)
        ril.addWidget(self.rpt_impaye_table)
        self.lbl_total_impaye = QLabel("Total impayé : 0.00")
        self.lbl_total_impaye.setStyleSheet(
            "font-weight: bold; color: #dc2626; padding: 5px;")
        ril.addWidget(self.lbl_total_impaye)
        self.report_tabs.addTab(ri, "Impayes")

        # Rapport vendeurs
        rvd = QWidget()
        rvdl = QVBoxLayout(rvd)
        rvdl.setContentsMargins(0, 8, 0, 0)
        self.rpt_vendeur_table = QTableWidget()
        self.rpt_vendeur_table.setColumnCount(5)
        self.rpt_vendeur_table.setHorizontalHeaderLabels(
            ["Vendeur", "Nb ventes", "CA total",
             "Montant payé", "Taux encaissement"])
        self._style_table(self.rpt_vendeur_table)
        rvdl.addWidget(self.rpt_vendeur_table)
        self.report_tabs.addTab(rvd, "Ventes par vendeur")

        report_layout.addWidget(self.report_tabs)
        layout.addWidget(report_frame)

        # Boutons export
        export_layout = QHBoxLayout()
        export_layout.addStretch()
        btn_pdf = QPushButton("Exporter PDF")
        btn_pdf.setFixedHeight(36)
        btn_pdf.setFixedWidth(140)
        btn_pdf.setObjectName("btn_pdf")
        btn_pdf.clicked.connect(self._export_pdf)
        btn_excel = QPushButton("Exporter Excel")
        btn_excel.setFixedHeight(36)
        btn_excel.setFixedWidth(140)
        btn_excel.setObjectName("btn_excel")
        btn_excel.clicked.connect(self._export_excel)
        export_layout.addWidget(btn_pdf)
        export_layout.addWidget(btn_excel)
        layout.addLayout(export_layout)

    # ----------------------------------------------------------
    # Helpers
    # ----------------------------------------------------------

    def _make_card(self, titre, valeur, color):
        frame = QFrame()
        frame.setObjectName("kpi_card")
        frame.setSizePolicy(QSizePolicy.Policy.Expanding,
                            QSizePolicy.Policy.Fixed)
        frame.setFixedHeight(100)
        layout = QVBoxLayout(frame)
        layout.setContentsMargins(15, 10, 15, 10)
        layout.setSpacing(3)
        bar = QFrame()
        bar.setFixedHeight(4)
        bar.setStyleSheet(f"background-color: {color}; border-radius: 2px;")
        layout.addWidget(bar)
        lbl_t = QLabel(titre)
        lbl_t.setStyleSheet("color: #64748b; font-size: 11px;")
        layout.addWidget(lbl_t)
        lbl_v = QLabel(valeur)
        lbl_v.setFont(QFont("Arial", 22, QFont.Weight.Bold))
        lbl_v.setStyleSheet(f"color: {color};")
        layout.addWidget(lbl_v)
        return {'frame': frame, 'valeur': lbl_v}

    def _make_mini_table(self, headers):
        table = QTableWidget()
        table.setColumnCount(len(headers))
        table.setHorizontalHeaderLabels(headers)
        self._style_table(table)
        return table

    def _style_table(self, table):
        table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        table.setAlternatingRowColors(True)
        table.verticalHeader().setVisible(False)
        table.horizontalHeader().setStretchLastSection(True)

    # ----------------------------------------------------------
    # Chargement donnees via controllers
    # ----------------------------------------------------------

    def refresh(self):
        self.lbl_refresh.setText(
            f"Mise à jour : {datetime.now().strftime('%H:%M:%S')}")
        self._load_ca()
        self._load_top_products()
        self._load_top_clients()
        self._load_stock_critique()
        self._load_graph()
        self._load_vendeurs_filter()
        self._load_report_data()

    def _load_ca(self):
        kpi = SaleController.get_kpi()
        self.card_jour['valeur'].setText(
            f"{float(kpi.get('ca_jour', 0)):.2f}")
        self.card_semaine['valeur'].setText(
            f"{float(kpi.get('ca_semaine', 0)):.2f}")
        self.card_mois['valeur'].setText(
            f"{float(kpi.get('ca_mois', 0)):.2f}")

    def _load_top_products(self):
        rows = SaleController.get_top_products(5)
        self.top_products_table.setRowCount(0)
        for i, r in enumerate(rows or []):
            self.top_products_table.insertRow(i)
            vals = [r['nom'], str(r['total_vendu']),
                    f"{float(r['ca']):.2f}"]
            for col, val in enumerate(vals):
                item = QTableWidgetItem(val)
                item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                if col == 1:
                    item.setForeground(QColor('#1976D2'))
                self.top_products_table.setItem(i, col, item)
        self.top_products_table.resizeColumnsToContents()

    def _load_top_clients(self):
        rows = SaleController.get_top_clients(5)
        self.top_clients_table.setRowCount(0)
        for i, r in enumerate(rows or []):
            self.top_clients_table.insertRow(i)
            vals = [r['client'], str(r['nb_ventes']),
                    f"{float(r['ca_total']):.2f}"]
            for col, val in enumerate(vals):
                item = QTableWidgetItem(val)
                item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                self.top_clients_table.setItem(i, col, item)
        self.top_clients_table.resizeColumnsToContents()

    def _load_stock_critique(self):
        rows = ProductController.get_low_stock()
        self.stock_table.setRowCount(0)
        if not rows:
            self.stock_table.insertRow(0)
            item = QTableWidgetItem("Aucun stock critique")
            item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            item.setForeground(QColor('#16a34a'))
            self.stock_table.setItem(0, 0, item)
            self.stock_table.setSpan(0, 0, 1, 3)
            return
        for i, r in enumerate(rows):
            self.stock_table.insertRow(i)
            vals = [r['nom'], str(r['stock_actuel']),
                    str(r['stock_min'])]
            for col, val in enumerate(vals):
                item = QTableWidgetItem(val)
                item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                if col == 1:
                    item.setForeground(QColor('#dc2626'))
                    f = item.font()
                    f.setBold(True)
                    item.setFont(f)
                self.stock_table.setItem(i, col, item)
        self.stock_table.resizeColumnsToContents()

    def _load_graph(self):
        rows = SaleController.get_monthly_revenue()
        if rows:
            labels = [r['mois'] for r in rows]
            values = [float(r['ca']) for r in rows]
            self.graph_widget.set_data(labels, values)

    def _load_vendeurs_filter(self):
        vendeurs = UserController.get_vendeurs()
        self.filter_vendeur.blockSignals(True)
        self.filter_vendeur.clear()
        self.filter_vendeur.addItem("Tous les vendeurs", None)
        for v in (vendeurs or []):
            self.filter_vendeur.addItem(v['username'], v['id'])
        self.filter_vendeur.blockSignals(False)

    # ----------------------------------------------------------
    # Rapports
    # ----------------------------------------------------------

    def _load_report_data(self):
        debut = self.date_debut.date().toString("yyyy-MM-dd")
        fin = self.date_fin.date().toString("yyyy-MM-dd")
        vendeur_id = self.filter_vendeur.currentData()
        self._load_rpt_ventes(debut, fin, vendeur_id)
        self._load_rpt_impayes(debut, fin, vendeur_id)
        self._load_rpt_vendeurs(debut, fin)

    def _load_rpt_ventes(self, debut, fin, vendeur_id):
        rows = SaleController.get_report_ventes(debut, fin, vendeur_id)
        self.rpt_ventes_table.setRowCount(0)
        total_ca, nb = 0, 0
        statut_colors = {'payee': '#16a34a', 'partielle': '#d97706',
                         'en_cours': '#1976D2'}
        for i, r in enumerate(rows):
            self.rpt_ventes_table.insertRow(i)
            statut = r.get('statut', '')
            vals = [r['numero_facture'], str(r['date_v']), r['client'],
                    r['vendeur'], f"{float(r['montant_total']):.2f}",
                    f"{float(r['montant_paye']):.2f}", statut.upper()]
            for col, val in enumerate(vals):
                item = QTableWidgetItem(val)
                item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                if col == 6:
                    item.setForeground(
                        QColor(statut_colors.get(statut, '#64748b')))
                    f = item.font()
                    f.setBold(True)
                    item.setFont(f)
                self.rpt_ventes_table.setItem(i, col, item)
            total_ca += float(r['montant_total'])
            nb += 1
        self.rpt_ventes_table.resizeColumnsToContents()
        self.lbl_total_ventes.setText(
            f"Total CA : {total_ca:.2f}  |  Nombre de ventes : {nb}")

    def _load_rpt_impayes(self, debut, fin, vendeur_id):
        rows = SaleController.get_report_impayes(debut, fin, vendeur_id)
        self.rpt_impaye_table.setRowCount(0)
        total_reste = 0
        for i, r in enumerate(rows):
            self.rpt_impaye_table.insertRow(i)
            vals = [r['numero_facture'], str(r['date_v']), r['client'],
                    f"{float(r['montant_total']):.2f}",
                    f"{float(r['montant_paye']):.2f}",
                    f"{float(r['montant_reste']):.2f}"]
            for col, val in enumerate(vals):
                item = QTableWidgetItem(val)
                item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                if col == 5:
                    item.setForeground(QColor('#dc2626'))
                    f = item.font()
                    f.setBold(True)
                    item.setFont(f)
                self.rpt_impaye_table.setItem(i, col, item)
            total_reste += float(r['montant_reste'])
        self.rpt_impaye_table.resizeColumnsToContents()
        self.lbl_total_impaye.setText(f"Total impayé : {total_reste:.2f}")

    def _load_rpt_vendeurs(self, debut, fin):
        rows = SaleController.get_report_vendeurs(debut, fin)
        self.rpt_vendeur_table.setRowCount(0)
        for i, r in enumerate(rows):
            self.rpt_vendeur_table.insertRow(i)
            ca = float(r['ca'])
            paye = float(r['paye'])
            taux = f"{(paye / ca * 100):.1f} %" if ca > 0 else "0 %"
            vals = [r['vendeur'], str(r['nb_ventes']),
                    f"{ca:.2f}", f"{paye:.2f}", taux]
            for col, val in enumerate(vals):
                item = QTableWidgetItem(val)
                item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                self.rpt_vendeur_table.setItem(i, col, item)
        self.rpt_vendeur_table.resizeColumnsToContents()

    # ----------------------------------------------------------
    # Exports
    # ----------------------------------------------------------

    def _get_current_report_data(self):
        idx = self.report_tabs.currentIndex()
        names = ['ventes_periode', 'impayes', 'ventes_vendeur']
        tables = [self.rpt_ventes_table,
                  self.rpt_impaye_table,
                  self.rpt_vendeur_table]
        return names[idx], tables[idx]

    def _export_pdf(self):
        name, table = self._get_current_report_data()
        debut = self.date_debut.date().toString("dd/MM/yyyy")
        fin = self.date_fin.date().toString("dd/MM/yyyy")
        filepath, _ = QFileDialog.getSaveFileName(
            self, "Enregistrer le rapport PDF",
            f"rapport_{name}_{datetime.now().strftime('%Y%m%d')}.pdf",
            "PDF (*.pdf)")
        if not filepath:
            return
        try:
            self._generate_report_pdf(filepath, name, table, debut, fin)
            QMessageBox.information(self, "Succes",
                                    f"Rapport PDF genere :\n{filepath}")
        except Exception as e:
            QMessageBox.critical(self, "Erreur", str(e))

    def _export_excel(self):
        name, table = self._get_current_report_data()
        filepath, _ = QFileDialog.getSaveFileName(
            self, "Enregistrer le rapport Excel",
            f"rapport_{name}_{datetime.now().strftime('%Y%m%d')}.xlsx",
            "Excel (*.xlsx)")
        if not filepath:
            return
        try:
            self._generate_report_excel(filepath, table)
            QMessageBox.information(self, "Succes",
                                    f"Rapport Excel genere :\n{filepath}")
        except Exception as e:
            QMessageBox.critical(self, "Erreur", str(e))

    def _generate_report_pdf(self, filepath, name, table, debut, fin):
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.platypus import (
            Paragraph,
            SimpleDocTemplate,
            Spacer,
            Table,
            TableStyle,
        )

        titles = {
            'ventes_periode': 'Rapport des Ventes par Période',
            'impayes': 'Rapport des Impayés',
            'ventes_vendeur': 'Rapport des Ventes par Vendeur',
        }
        DARK = colors.HexColor('#1e293b')
        BLUE = colors.HexColor('#1976D2')
        LIGHT = colors.HexColor('#f8fafc')

        doc = SimpleDocTemplate(filepath, pagesize=landscape(A4),
                                rightMargin=1.5 * cm, leftMargin=1.5 * cm,
                                topMargin=1.5 * cm, bottomMargin=1.5 * cm)
        s_title = ParagraphStyle('T', fontSize=14, fontName='Helvetica-Bold',
                                 textColor=DARK)
        s_sub = ParagraphStyle('S', fontSize=9, fontName='Helvetica',
                               textColor=colors.HexColor('#64748b'))
        story = []
        story.append(Paragraph(titles.get(name, 'Rapport'), s_title))
        story.append(Paragraph(f"Période : {debut} au {fin}", s_sub))
        story.append(Paragraph(
            f"Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}", s_sub))
        story.append(Spacer(1, 0.5 * cm))

        headers = [table.horizontalHeaderItem(c).text()
                   for c in range(table.columnCount())]
        data = [headers]
        for row in range(table.rowCount()):
            data.append([
                (table.item(row, col).text()
                 if table.item(row, col) else '')
                for col in range(table.columnCount())
            ])

        W = landscape(A4)[0] - 3 * cm
        col_w = [W / table.columnCount()] * table.columnCount()
        t = Table(data, colWidths=col_w, repeatRows=1)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), DARK),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('TOPPADDING', (0, 0), (-1, -1), 5),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, LIGHT]),
            ('GRID', (0, 0), (-1, -1), 0.3,
             colors.HexColor('#e2e8f0')),
            ('LINEBELOW', (0, 0), (-1, 0), 1.5, BLUE),
        ]))
        story.append(t)
        doc.build(story)

    def _generate_report_excel(self, filepath, table):
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Rapport"
        DARK_FILL = PatternFill("solid", fgColor="1e293b")
        LIGHT_FILL = PatternFill("solid", fgColor="f8fafc")

        headers = [table.horizontalHeaderItem(c).text()
                   for c in range(table.columnCount())]
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=ci, value=h)
            cell.fill = DARK_FILL
            cell.font = Font(color="FFFFFF", bold=True, size=10)
            cell.alignment = Alignment(horizontal='center',
                                       vertical='center')
        ws.row_dimensions[1].height = 22

        for ri in range(table.rowCount()):
            fill = LIGHT_FILL if ri % 2 == 0 else PatternFill()
            for ci in range(table.columnCount()):
                item = table.item(ri, ci)
                cell = ws.cell(row=ri + 2, column=ci + 1,
                               value=item.text() if item else '')
                cell.fill = fill
                cell.alignment = Alignment(horizontal='center')

        for col in ws.columns:
            max_len = max(
                (len(str(c.value)) for c in col if c.value), default=10)
            ws.column_dimensions[
                col[0].column_letter].width = min(max_len + 4, 35)
        wb.save(filepath)

    # ----------------------------------------------------------
    # Styles
    # ----------------------------------------------------------

    def _apply_styles(self):
        self.setStyleSheet("""
            QFrame#kpi_card, QFrame#stat_frame {
                background: white;
                border: 1px solid #e2e8f0;
                border-radius: 10px;
            }
            QPushButton#btn_primary {
                background-color: #1976D2; color: white;
            }
            QPushButton#btn_primary:hover { background-color: #1565C0; }
            QPushButton#btn_excel {
                background-color: #dcfce7; color: #16a34a;
            }
            QPushButton#btn_excel:hover { background-color: #bbf7d0; }
        """)


# ──────────────────────────────────────────────────────────────
# Widget graphique courbe (PyQt6 pur)
# ──────────────────────────────────────────────────────────────

class BarChartWidget(QWidget):
    """Graphique courbe evolution CA en PyQt6 pur."""

    def __init__(self, parent=None):
        super().__init__(parent)
        self.labels = []
        self.values = []
        self.setMinimumHeight(150)

    def set_data(self, labels, values):
        self.labels = labels
        self.values = values
        self.update()

    def paintEvent(self, event):
        from PyQt6.QtCore import QPointF
        from PyQt6.QtGui import (
            QBrush,
            QLinearGradient,
            QPainter,
            QPainterPath,
            QPen,
        )

        if not self.values:
            return

        painter = QPainter(self)
        painter.setRenderHint(QPainter.RenderHint.Antialiasing)

        W = self.width()
        H = self.height()
        pad_l, pad_r, pad_t, pad_b = 65, 25, 20, 35
        chart_w = W - pad_l - pad_r
        chart_h = H - pad_t - pad_b
        n = len(self.values)
        max_val = max(self.values) if max(self.values) > 0 else 1

        painter.fillRect(0, 0, W, H, QColor('white'))

        nb_grid = 4
        painter.setFont(QFont('Arial', 7))
        for i in range(nb_grid + 1):
            gy = pad_t + chart_h * i // nb_grid
            gv = max_val * (nb_grid - i) / nb_grid
            painter.setPen(QPen(QColor('#e2e8f0'), 1))
            painter.drawLine(pad_l, gy, W - pad_r, gy)
            painter.setPen(QColor('#94a3b8'))
            painter.drawText(0, gy - 6, pad_l - 5, 14,
                             Qt.AlignmentFlag.AlignRight, f"{gv:.0f}")

        def pt(i):
            x = (pad_l + i * chart_w / (n - 1)
                 if n > 1 else pad_l + chart_w / 2)
            y = (pad_t + chart_h
                 - self.values[i] / max_val * chart_h)
            return QPointF(x, y)

        points = [pt(i) for i in range(n)]

        fill_path = QPainterPath()
        fill_path.moveTo(QPointF(points[0].x(), pad_t + chart_h))
        for p in points:
            fill_path.lineTo(p)
        fill_path.lineTo(QPointF(points[-1].x(), pad_t + chart_h))
        fill_path.closeSubpath()

        grad = QLinearGradient(
            QPointF(0, pad_t), QPointF(0, pad_t + chart_h))
        grad.setColorAt(0.0, QColor(25, 118, 210, 80))
        grad.setColorAt(1.0, QColor(25, 118, 210, 5))
        painter.setBrush(QBrush(grad))
        painter.setPen(Qt.PenStyle.NoPen)
        painter.drawPath(fill_path)

        line_path = QPainterPath()
        line_path.moveTo(points[0])
        for i in range(1, n):
            prev = points[i - 1]
            curr = points[i]
            cx = (prev.x() + curr.x()) / 2
            line_path.cubicTo(
                QPointF(cx, prev.y()),
                QPointF(cx, curr.y()), curr)

        painter.setBrush(Qt.BrushStyle.NoBrush)
        painter.setPen(QPen(QColor('#1976D2'), 2.5))
        painter.drawPath(line_path)

        for i, (p, val) in enumerate(zip(points, self.values)):
            painter.setBrush(QBrush(QColor('white')))
            painter.setPen(QPen(QColor('#1976D2'), 2))
            painter.drawEllipse(p, 5, 5)
            painter.setPen(QColor('#1e293b'))
            painter.setFont(QFont('Arial', 7, QFont.Weight.Bold))
            painter.drawText(int(p.x()) - 25, int(p.y()) - 16, 50, 14,
                             Qt.AlignmentFlag.AlignCenter, f"{val:.0f}")
            painter.setPen(QColor('#64748b'))
            painter.setFont(QFont('Arial', 7))
            painter.drawText(int(p.x()) - 25, H - pad_b + 5, 50, 14,
                             Qt.AlignmentFlag.AlignCenter, self.labels[i])

        painter.end()
